# -*- coding: utf-8 -*-
import os
import re
import sys
import time
from difflib import SequenceMatcher
from urllib.parse import parse_qs, urlparse

import django
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


os.environ.setdefault("DJANGO_SETTINGS_MODULE", "HomePage.settings")
django.setup()

from myHomePage.services import GoogleScholarService  # noqa: E402


SCHOLAR_ID = "zwmacb4AAAAJ"
AUTHOR_NAME = "Xiaoheng Deng"
FILE_NAME = "dxc.xlsx"

COL_TITLE = "\u6807\u9898"
COL_FIRST = "\u662f\u5426\u4e00\u4f5c"
COL_CORRESPONDING = "\u662f\u5426\u901a\u4fe1"
COL_AUTHORS = "\u4f5c\u8005"
COL_JOURNAL = "\u671f\u520a"
COL_DATE = "\u53d1\u8868\u65f6\u95f4"
COL_VOLUME = "\u5377\u53f7"
COL_ISSUE = "\u671f\u53f7"
COL_PAGES = "\u9875\u7801"
COL_DOI = "DOI"
COL_CITATIONS = "\u5f15\u7528\u6570"
COL_LINK = "\u94fe\u63a5"
COL_BIBTEX = "BibTeX"

COLUMNS = [
    COL_TITLE,
    COL_FIRST,
    COL_CORRESPONDING,
    COL_AUTHORS,
    COL_JOURNAL,
    COL_DATE,
    COL_VOLUME,
    COL_ISSUE,
    COL_PAGES,
    COL_DOI,
    COL_CITATIONS,
    COL_LINK,
    COL_BIBTEX,
]

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}


def clean(value):
    return str(value or "").replace("\r\n", " ").replace("\n", " ").strip()


def normalize_title(value):
    value = clean(value).lower()
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return " ".join(value.split())


def normalize_name(value):
    value = re.sub(r"\([^)]*\)", "", clean(value))
    value = re.sub(r"[^a-zA-Z\s]", " ", value).lower()
    return " ".join(value.split())


def normalize_doi(value):
    value = clean(value).lower()
    value = value.replace("https://doi.org/", "").replace("http://doi.org/", "")
    value = value.replace("doi:", "").strip().rstrip(".,;")
    return value


def name_variants(full_name):
    parts = normalize_name(full_name).split()
    variants = {normalize_name(full_name)}
    if len(parts) >= 2:
        first, last = parts[0], parts[-1]
        variants.update({f"{first[0]} {last}", f"{last} {first}", f"{last} {first[0]}"})
    return variants


AUTHOR_VARIANTS = name_variants(AUTHOR_NAME)


def split_authors(authors):
    return [item.strip() for item in clean(authors).split(",") if item.strip()]


def is_first_author(authors):
    author_list = split_authors(authors)
    if not author_list:
        return "\u672a\u77e5"
    return "\u662f" if normalize_name(author_list[0]) in AUTHOR_VARIANTS else "\u5426"


def parse_venue(venue):
    venue = clean(venue).replace("\xa0", " ")
    result = {COL_JOURNAL: venue, COL_VOLUME: "", COL_ISSUE: "", COL_PAGES: ""}
    if not venue:
        return result

    # Examples:
    # IEEE Internet of Things Journal 10 (4), 2954-2966
    # Engineering Applications of Artificial Intelligence 181, 115309
    match = re.match(
        r"^(?P<journal>.*?)(?:\s+(?P<volume>\d+[A-Za-z]?)(?:\s*\((?P<issue>[^)]+)\))?)?(?:,\s*(?P<pages>[^,]+))?$",
        venue,
    )
    if match and match.group("journal"):
        groups = {key: clean(value) for key, value in match.groupdict().items()}
        result[COL_JOURNAL] = groups.get("journal") or venue
        result[COL_VOLUME] = groups.get("volume") or ""
        result[COL_ISSUE] = groups.get("issue") or ""
        result[COL_PAGES] = groups.get("pages") or ""
    return result


def scholar_params_from_url(url):
    query = parse_qs(urlparse(url).query)
    return {key: values[0] for key, values in query.items() if values}


def parse_scholar_page(html):
    soup = BeautifulSoup(html, "html.parser")
    records = []
    for item in soup.select("#gsc_a_b .gsc_a_tr"):
        title_element = item.select_one(".gsc_a_t a")
        if not title_element:
            continue
        gray = item.select(".gs_gray")
        authors = gray[0].get_text(" ", strip=True) if len(gray) > 0 else ""
        venue = gray[1].get_text(" ", strip=True) if len(gray) > 1 else ""
        year_element = item.select_one(".gsc_a_y")
        citation_element = item.select_one(".gsc_a_c")
        href = title_element.get("href") or ""
        url = f"https://scholar.google.com{href}" if href.startswith("/") else href
        year = clean(year_element.get_text(" ", strip=True) if year_element else "")
        citations_text = clean(citation_element.get_text(" ", strip=True) if citation_element else "")
        venue_parts = parse_venue(venue)
        records.append({
            COL_TITLE: clean(title_element.get_text(" ", strip=True)),
            COL_FIRST: is_first_author(authors),
            COL_CORRESPONDING: "\u672a\u77e5\uff08Google Scholar\u672a\u63d0\u4f9b\uff09",
            COL_AUTHORS: authors,
            COL_JOURNAL: venue_parts[COL_JOURNAL],
            COL_DATE: year,
            COL_VOLUME: venue_parts[COL_VOLUME],
            COL_ISSUE: venue_parts[COL_ISSUE],
            COL_PAGES: venue_parts[COL_PAGES],
            COL_DOI: "",
            COL_CITATIONS: int(citations_text) if citations_text.isdigit() else 0,
            COL_LINK: url,
            COL_BIBTEX: "",
            "_scholar_url": url,
        })
    return records


def fetch_scholar_pubdate_records():
    service = GoogleScholarService(SCHOLAR_ID)
    records = []
    seen = set()
    for start in range(0, 1000, 100):
        params = {
            "user": SCHOLAR_ID,
            "hl": "en",
            "cstart": start,
            "pagesize": 100,
            "sortby": "pubdate",
        }
        response = service._request_scholar(params, HEADERS)
        if "sorry/index" in response.url or "not a robot" in response.text.lower():
            raise RuntimeError("Google Scholar returned a verification page")
        batch = parse_scholar_page(response.text)
        print(f"scholar start={start}: {len(batch)}")
        if not batch:
            break
        for record in batch:
            key = normalize_title(record[COL_TITLE])
            if key and key not in seen:
                seen.add(key)
                records.append(record)
        if len(batch) < 100:
            break
        time.sleep(0.3)
    return records


def load_existing_records():
    workbook = load_workbook(FILE_NAME)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    records = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        record = {header: value for header, value in zip(headers, row)}
        records.append(record)
    return workbook, sheet, records


def title_similarity(a, b):
    return SequenceMatcher(None, normalize_title(a), normalize_title(b)).ratio()


def first_list_value(value):
    if isinstance(value, list):
        return clean(value[0]) if value else ""
    return clean(value)


def first_date_part(message):
    for key in ("published-print", "published-online", "published", "issued"):
        parts = (message.get(key) or {}).get("date-parts") or []
        if parts and parts[0]:
            return "-".join(str(part).zfill(2) for part in parts[0])
    return ""


def crossref_by_title(session, title, year):
    response = session.get(
        "https://api.crossref.org/works",
        params={"query.title": title, "rows": 3},
        headers={"User-Agent": "HomePageScholarExport/1.0"},
        timeout=20,
    )
    response.raise_for_status()
    best = None
    best_score = 0
    for candidate in response.json().get("message", {}).get("items", []):
        candidate_title = " ".join(candidate.get("title") or [])
        score = title_similarity(title, candidate_title)
        if year:
            issued = (candidate.get("issued") or {}).get("date-parts") or []
            candidate_year = issued[0][0] if issued and issued[0] else None
            if candidate_year and abs(int(candidate_year) - int(year)) > 1:
                score -= 0.2
        if score > best_score:
            best = candidate
            best_score = score
    return best if best_score >= 0.9 else None


def apply_crossref(record, message):
    if not message:
        return False
    changed = False
    mapping = {
        COL_DOI: normalize_doi(message.get("DOI")),
        COL_JOURNAL: first_list_value(message.get("container-title")),
        COL_VOLUME: clean(message.get("volume")),
        COL_ISSUE: clean(message.get("issue")),
        COL_PAGES: clean(message.get("page")),
        COL_DATE: first_date_part(message),
    }
    for key, value in mapping.items():
        if value and (not clean(record.get(key)) or key == COL_DOI):
            record[key] = value
            changed = True
    if mapping[COL_DOI]:
        record[COL_LINK] = f"https://doi.org/{mapping[COL_DOI]}"
    return changed


def bibtex_key(title, year):
    words = re.findall(r"[A-Za-z0-9]+", title or "")
    return "deng{}{}".format(year or "", ("".join(words[:5]) or "publication")[:60])


def year_from_date(value):
    match = re.search(r"\d{4}", clean(value))
    return match.group(0) if match else ""


def make_bibtex(record):
    year = year_from_date(record.get(COL_DATE))
    journal = clean(record.get(COL_JOURNAL))
    entry_type = "misc" if journal.lower() in {"corr", "arxiv"} else "article"
    fields = [
        ("title", record.get(COL_TITLE)),
        ("author", clean(record.get(COL_AUTHORS)).replace(", ", " and ")),
        ("journal", journal),
        ("year", year),
        ("volume", record.get(COL_VOLUME)),
        ("number", record.get(COL_ISSUE)),
        ("pages", record.get(COL_PAGES)),
        ("doi", record.get(COL_DOI)),
        ("url", record.get(COL_LINK)),
    ]
    lines = [f"@{entry_type}{{{bibtex_key(record.get(COL_TITLE), year)},"]
    for key, value in fields:
        value = clean(value)
        if value and value.lower() != "none":
            lines.append(f"  {key} = {{{value}}},")
    if len(lines) > 1:
        lines[-1] = lines[-1].rstrip(",")
    lines.append("}")
    return "\n".join(lines)


def merge_records(existing, scholar_records):
    by_title = {normalize_title(record.get(COL_TITLE)): record for record in existing if normalize_title(record.get(COL_TITLE))}
    added = 0
    updated = 0
    for scholar_record in scholar_records:
        key = normalize_title(scholar_record.get(COL_TITLE))
        if not key:
            continue
        if key not in by_title:
            by_title[key] = {column: scholar_record.get(column, "") for column in COLUMNS}
            added += 1
            continue
        record = by_title[key]
        for column in COLUMNS:
            if column == COL_BIBTEX:
                continue
            incoming = scholar_record.get(column, "")
            if incoming and not clean(record.get(column)):
                record[column] = incoming
                updated += 1
        # Scholar has the freshest citation counts and sorted publication years.
        if scholar_record.get(COL_CITATIONS) not in ("", None):
            record[COL_CITATIONS] = scholar_record.get(COL_CITATIONS)
    return list(by_title.values()), added, updated


def enrich_newer_records(records):
    session = requests.Session()
    session.trust_env = False
    enriched = 0
    for index, record in enumerate(records, start=1):
        year = year_from_date(record.get(COL_DATE))
        if year and int(year) < 2024 and clean(record.get(COL_DOI)):
            record[COL_BIBTEX] = make_bibtex(record)
            continue
        if clean(record.get(COL_DOI)) and clean(record.get(COL_VOLUME)) and clean(record.get(COL_PAGES)):
            record[COL_BIBTEX] = make_bibtex(record)
            continue
        try:
            message = crossref_by_title(session, record.get(COL_TITLE), year)
            if apply_crossref(record, message):
                enriched += 1
        except Exception as exc:
            print(f"crossref failed {index}: {clean(record.get(COL_TITLE))[:80]} ({exc})")
        record[COL_BIBTEX] = make_bibtex(record)
        time.sleep(0.05)
    return enriched


def write_records(sheet, records):
    sheet.delete_rows(1, sheet.max_row)
    sheet.append(COLUMNS)
    records.sort(key=lambda item: (year_from_date(item.get(COL_DATE)) or "0000", clean(item.get(COL_DATE)), int(item.get(COL_CITATIONS) or 0)), reverse=True)
    for record in records:
        sheet.append([record.get(column, "") for column in COLUMNS])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    widths = [56, 10, 24, 44, 36, 16, 10, 10, 18, 32, 10, 48, 76]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def main():
    do_enrich = "--enrich" in sys.argv
    workbook, sheet, existing = load_existing_records()
    scholar_records = fetch_scholar_pubdate_records()
    merged, added, updated = merge_records(existing, scholar_records)
    if do_enrich:
        enriched = enrich_newer_records(merged)
    else:
        enriched = 0
        for record in merged:
            record[COL_BIBTEX] = make_bibtex(record)
    write_records(sheet, merged)
    workbook.save(FILE_NAME)
    print(
        f"scholar={len(scholar_records)} existing={len(existing)} added={added} "
        f"updated_fields={updated} enriched={enriched} total={len(merged)}"
    )


if __name__ == "__main__":
    main()
