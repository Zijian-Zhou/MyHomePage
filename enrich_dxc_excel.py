# -*- coding: utf-8 -*-
import re
import time
from difflib import SequenceMatcher
from urllib.parse import quote

import requests
from openpyxl import load_workbook


FILE_NAME = "dxc.xlsx"
REQUEST_DELAY_SECONDS = 0.05

COL_TITLE = "\u6807\u9898"
COL_FIRST = "\u662f\u5426\u4e00\u4f5c"
COL_CORRESPONDING = "\u662f\u5426\u901a\u4fe1"
COL_AUTHORS = "\u4f5c\u8005"
COL_JOURNAL = "\u671f\u520a"
COL_DATE = "\u53d1\u8868\u65f6\u95f4"
COL_YEAR = "\u5e74\u4efd"
COL_VOLUME = "\u5377\u53f7"
COL_ISSUE = "\u671f\u53f7"
COL_PAGES = "\u9875\u7801"
COL_DOI = "DOI"
COL_CITATIONS = "\u5f15\u7528\u6570"
COL_LINK = "\u94fe\u63a5"
COL_BIBTEX = "BibTeX"


def clean(value):
    return str(value or "").replace("\r\n", " ").replace("\n", " ").strip()


def normalize_title(value):
    value = clean(value).lower()
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return " ".join(value.split())


def title_similarity(a, b):
    return SequenceMatcher(None, normalize_title(a), normalize_title(b)).ratio()


def normalize_doi(value):
    value = clean(value).lower()
    value = value.replace("https://doi.org/", "").replace("http://doi.org/", "")
    value = value.replace("doi:", "").strip().rstrip(".,;")
    return value


def first_date_part(message):
    for key in ("published-print", "published-online", "published", "issued"):
        parts = (message.get(key) or {}).get("date-parts") or []
        if parts and parts[0]:
            return "-".join(str(part).zfill(2) for part in parts[0])
    return ""


def first_list_value(value):
    if isinstance(value, list):
        return clean(value[0]) if value else ""
    return clean(value)


def crossref_work_by_doi(session, doi):
    if not doi:
        return None
    response = session.get(
        f"https://api.crossref.org/works/{quote(doi, safe='')}",
        headers={"User-Agent": "HomePageScholarExport/1.0"},
        timeout=25,
    )
    if response.status_code == 404:
        return None
    response.raise_for_status()
    return response.json().get("message")


def crossref_work_by_title(session, title, year):
    response = session.get(
        "https://api.crossref.org/works",
        params={"query.title": title, "rows": 5},
        headers={"User-Agent": "HomePageScholarExport/1.0"},
        timeout=25,
    )
    response.raise_for_status()
    candidates = response.json().get("message", {}).get("items", [])
    best = None
    best_score = 0
    for candidate in candidates:
        candidate_title = " ".join(candidate.get("title") or [])
        score = title_similarity(title, candidate_title)
        if year:
            candidate_year = None
            issued = (candidate.get("issued") or {}).get("date-parts") or []
            if issued and issued[0]:
                candidate_year = issued[0][0]
            if candidate_year and abs(int(candidate_year) - int(year)) > 1:
                score -= 0.1
        if score > best_score:
            best = candidate
            best_score = score
    return best if best_score >= 0.9 else None


def openalex_work_by_doi(session, doi):
    if not doi:
        return None
    response = session.get(
        f"https://api.openalex.org/works/https://doi.org/{doi}",
        timeout=25,
    )
    if response.status_code == 404:
        return None
    response.raise_for_status()
    return response.json()


def apply_crossref(record, message):
    if not message:
        return False
    changed = False
    mapping = {
        COL_DOI: normalize_doi(message.get("DOI")),
        COL_JOURNAL: first_list_value(message.get("container-title")),
        COL_VOLUME: clean(message.get("volume")),
        COL_ISSUE: clean(message.get("issue")),
        COL_PAGES: clean(message.get("page")),
        COL_DATE: first_date_part(message),
    }
    for key, value in mapping.items():
        if value and not clean(record.get(key)):
            record[key] = value
            changed = True
    if mapping[COL_DOI]:
        if not clean(record.get(COL_DOI)):
            changed = True
        record[COL_DOI] = mapping[COL_DOI]
        record[COL_LINK] = f"https://doi.org/{mapping[COL_DOI]}"
    return changed


def apply_openalex(record, message):
    if not message:
        return False
    changed = False
    biblio = message.get("biblio") or {}
    primary = message.get("primary_location") or {}
    source = primary.get("source") or {}
    pages = clean(biblio.get("first_page"))
    if biblio.get("last_page"):
        pages = f"{pages}-{biblio.get('last_page')}" if pages else clean(biblio.get("last_page"))
    mapping = {
        COL_JOURNAL: clean(source.get("display_name")),
        COL_VOLUME: clean(biblio.get("volume")),
        COL_ISSUE: clean(biblio.get("issue")),
        COL_PAGES: pages,
        COL_DATE: clean(message.get("publication_date")),
    }
    doi = normalize_doi(message.get("doi"))
    if doi:
        mapping[COL_DOI] = doi
    for key, value in mapping.items():
        if value and not clean(record.get(key)):
            record[key] = value
            changed = True
    if doi and not clean(record.get(COL_LINK)):
        record[COL_LINK] = f"https://doi.org/{doi}"
        changed = True
    return changed


def bibtex_key(title, year):
    words = re.findall(r"[A-Za-z0-9]+", title or "")
    return "deng{}{}".format(year or "", ("".join(words[:5]) or "publication")[:60])


def make_bibtex(record):
    fields = [
        ("title", record.get(COL_TITLE)),
        ("author", clean(record.get(COL_AUTHORS)).replace(", ", " and ")),
        ("journal", record.get(COL_JOURNAL)),
        ("year", record.get(COL_YEAR)),
        ("volume", record.get(COL_VOLUME)),
        ("number", record.get(COL_ISSUE)),
        ("pages", record.get(COL_PAGES)),
        ("doi", record.get(COL_DOI)),
        ("url", record.get(COL_LINK)),
    ]
    lines = [f"@article{{{bibtex_key(record.get(COL_TITLE), record.get(COL_YEAR))},"]
    for key, value in fields:
        value = clean(value)
        if value:
            lines.append(f"  {key} = {{{value}}},")
    if len(lines) > 1:
        lines[-1] = lines[-1].rstrip(",")
    lines.append("}")
    return "\n".join(lines)


def missing_counts(records, columns):
    return {
        column: sum(1 for record in records if not clean(record.get(column)))
        for column in columns
    }


def main():
    workbook = load_workbook(FILE_NAME)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    records = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        records.append({header: value for header, value in zip(headers, row)})

    tracked = [COL_JOURNAL, COL_VOLUME, COL_ISSUE, COL_PAGES, COL_DOI, COL_LINK, COL_BIBTEX]
    print("before", missing_counts(records, tracked))

    session = requests.Session()
    session.trust_env = False
    updated = 0
    for index, record in enumerate(records, start=1):
        changed = False
        doi = normalize_doi(record.get(COL_DOI))
        title = clean(record.get(COL_TITLE))
        year = record.get(COL_YEAR)

        try:
            message = crossref_work_by_doi(session, doi) if doi else None
            if not message:
                message = crossref_work_by_title(session, title, year)
            changed = apply_crossref(record, message) or changed
        except Exception as exc:
            print(f"crossref failed {index}: {title[:80]} ({exc})")

        doi = normalize_doi(record.get(COL_DOI))
        if doi and (not clean(record.get(COL_VOLUME)) or not clean(record.get(COL_ISSUE)) or not clean(record.get(COL_PAGES))):
            try:
                changed = apply_openalex(record, openalex_work_by_doi(session, doi)) or changed
            except Exception as exc:
                print(f"openalex failed {index}: {title[:80]} ({exc})")

        record[COL_BIBTEX] = make_bibtex(record)
        if changed:
            updated += 1
        time.sleep(REQUEST_DELAY_SECONDS)

    for row_index, record in enumerate(records, start=2):
        for column_index, header in enumerate(headers, start=1):
            sheet.cell(row=row_index, column=column_index).value = record.get(header, "")

    workbook.save(FILE_NAME)
    print("updated", updated)
    print("after", missing_counts(records, tracked))


if __name__ == "__main__":
    main()
