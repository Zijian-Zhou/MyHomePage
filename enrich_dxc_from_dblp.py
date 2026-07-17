# -*- coding: utf-8 -*-
import re
import time
from difflib import SequenceMatcher

import requests
from openpyxl import load_workbook


FILE_NAME = "dxc.xlsx"

COL_TITLE = "\u6807\u9898"
COL_JOURNAL = "\u671f\u520a"
COL_DATE = "\u53d1\u8868\u65f6\u95f4"
COL_VOLUME = "\u5377\u53f7"
COL_ISSUE = "\u671f\u53f7"
COL_PAGES = "\u9875\u7801"
COL_DOI = "DOI"
COL_LINK = "\u94fe\u63a5"
COL_AUTHORS = "\u4f5c\u8005"
COL_BIBTEX = "BibTeX"


def clean(value):
    return str(value or "").replace("\r\n", " ").replace("\n", " ").strip()


def normalize_title(value):
    value = clean(value).lower()
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return " ".join(value.split())


def similarity(a, b):
    return SequenceMatcher(None, normalize_title(a), normalize_title(b)).ratio()


def search_dblp(session, title):
    response = session.get(
        "https://dblp.org/search/publ/api",
        params={"q": title, "format": "json"},
        timeout=25,
    )
    response.raise_for_status()
    hits = response.json().get("result", {}).get("hits", {}).get("hit", [])
    if isinstance(hits, dict):
        hits = [hits]
    best = None
    best_score = 0
    for hit in hits:
        info = hit.get("info") or {}
        score = similarity(title, info.get("title"))
        if score > best_score:
            best = info
            best_score = score
    return best if best_score >= 0.9 else None


def extract_year(date_value):
    match = re.search(r"\d{4}", clean(date_value))
    return match.group(0) if match else ""


def bibtex_key(title, year):
    words = re.findall(r"[A-Za-z0-9]+", title or "")
    return "deng{}{}".format(year or "", ("".join(words[:5]) or "publication")[:60])


def make_bibtex(record):
    year = extract_year(record.get(COL_DATE))
    entry_type = "misc" if clean(record.get(COL_JOURNAL)).lower() in {"corr", "arxiv"} else "article"
    fields = [
        ("title", record.get(COL_TITLE)),
        ("author", clean(record.get(COL_AUTHORS)).replace(", ", " and ")),
        ("journal", record.get(COL_JOURNAL)),
        ("year", year),
        ("volume", record.get(COL_VOLUME)),
        ("number", record.get(COL_ISSUE)),
        ("pages", record.get(COL_PAGES)),
        ("doi", record.get(COL_DOI)),
        ("url", record.get(COL_LINK)),
    ]
    lines = [f"@{entry_type}{{{bibtex_key(record.get(COL_TITLE), year)},"]
    for key, value in fields:
        value = clean(value)
        if value:
            lines.append(f"  {key} = {{{value}}},")
    if len(lines) > 1:
        lines[-1] = lines[-1].rstrip(",")
    lines.append("}")
    return "\n".join(lines)


def main():
    workbook = load_workbook(FILE_NAME)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    records = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        records.append({header: value for header, value in zip(headers, row)})

    session = requests.Session()
    session.trust_env = False
    updated = 0
    for index, record in enumerate(records, start=1):
        if clean(record.get(COL_VOLUME)) and clean(record.get(COL_ISSUE)) and clean(record.get(COL_PAGES)):
            continue
        title = clean(record.get(COL_TITLE))
        try:
            info = search_dblp(session, title)
        except Exception as exc:
            print(f"dblp failed {index}: {title[:80]} ({exc})")
            continue
        if not info:
            continue
        changed = False
        mapping = {
            COL_JOURNAL: info.get("venue"),
            COL_VOLUME: info.get("volume"),
            COL_ISSUE: info.get("number"),
            COL_PAGES: info.get("pages"),
            COL_LINK: info.get("ee") or info.get("url"),
        }
        for key, value in mapping.items():
            value = clean(value)
            if value and not clean(record.get(key)):
                record[key] = value
                changed = True
        if changed:
            record[COL_BIBTEX] = make_bibtex(record)
            updated += 1
        time.sleep(0.05)

    for row_index, record in enumerate(records, start=2):
        for column_index, header in enumerate(headers, start=1):
            sheet.cell(row=row_index, column=column_index).value = record.get(header, "")
    workbook.save(FILE_NAME)
    print("updated", updated)


if __name__ == "__main__":
    main()
