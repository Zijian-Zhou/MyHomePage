import os
import re
import time
from datetime import datetime
from urllib.parse import parse_qs, urlparse

import django
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


os.environ.setdefault("DJANGO_SETTINGS_MODULE", "HomePage.settings")
django.setup()

from myHomePage.services import (  # noqa: E402
    GoogleScholarService,
    _entry_to_bibtex,
    _normalize_doi,
    _sanitize_bibtex_value,
)


SCHOLAR_ID = "zwmacb4AAAAJ"
PROFILE_NAME = "Xiaoheng Deng"
OUTPUT_FILE = "dxc.xlsx"
PAGE_SIZE = 100
REQUEST_DELAY_SECONDS = 0.25

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}


def normalize_name(value):
    value = re.sub(r"\([^)]*\)", "", value or "")
    value = re.sub(r"[^a-zA-Z\s]", " ", value).lower()
    return " ".join(value.split())


def name_variants(full_name):
    normalized = normalize_name(full_name)
    parts = normalized.split()
    variants = {normalized}
    if len(parts) >= 2:
        first, last = parts[0], parts[-1]
        variants.add(f"{first[0]} {last}")
        variants.add(f"{last} {first}")
        variants.add(f"{last} {first[0]}")
    return variants


PROFILE_NAME_VARIANTS = name_variants(PROFILE_NAME)


def split_authors(authors):
    if not authors:
        return []
    if " and " in authors:
        return [a.strip() for a in authors.split(" and ") if a.strip()]
    return [a.strip() for a in authors.split(",") if a.strip()]


def is_first_author(authors):
    author_list = split_authors(authors)
    if not author_list:
        return "未知"
    first = normalize_name(author_list[0])
    return "是" if first in PROFILE_NAME_VARIANTS else "否"


def scholar_params_from_url(url):
    query = parse_qs(urlparse(url).query)
    return {key: values[0] for key, values in query.items() if values}


def fetch_all_publications(service):
    all_publications = []
    seen_titles = set()
    start = 0

    while True:
        batch = service.get_publications(start=start, count=PAGE_SIZE)
        if not batch:
            break

        new_count = 0
        for pub in batch:
            title_key = (pub.get("title") or "").strip().lower()
            if not title_key or title_key in seen_titles:
                continue
            seen_titles.add(title_key)
            all_publications.append(pub)
            new_count += 1

        if len(batch) < PAGE_SIZE or new_count == 0:
            break
        start += PAGE_SIZE
        time.sleep(REQUEST_DELAY_SECONDS)

    return all_publications


def fetch_detail(service, pub):
    params = scholar_params_from_url(pub.get("url", ""))
    if not params:
        return {}

    response = service._request_scholar(params, HEADERS)
    soup = BeautifulSoup(response.text, "html.parser")

    details = {}
    title = soup.select_one("#gsc_oci_title")
    if title:
        details["Title"] = title.get_text(" ", strip=True)

    for row in soup.select("#gsc_oci_table .gs_scl"):
        field = row.select_one(".gsc_oci_field")
        value = row.select_one(".gsc_oci_value")
        if field and value:
            details[field.get_text(" ", strip=True)] = value.get_text(" ", strip=True)

    return details


def extract_doi_from_text(*values):
    doi_pattern = re.compile(r"10\.\d{4,9}/[-._;()/:A-Z0-9]+", re.IGNORECASE)
    for value in values:
        if not value:
            continue
        match = doi_pattern.search(str(value))
        if match:
            return _normalize_doi(match.group(0).rstrip(".,;"))
    return None


def normalize_title(title):
    title = (title or "").lower()
    title = re.sub(r"[^a-z0-9]+", " ", title)
    return " ".join(title.split())


def crossref_doi(title, year=None):
    try:
        response = requests.get(
            "https://api.crossref.org/works",
            params={"query.title": title, "rows": 1, "select": "DOI,title,issued,published-print,published-online"},
            headers={"User-Agent": "HomePageScholarExport/1.0 (mailto:unknown@example.com)"},
            timeout=20,
        )
        response.raise_for_status()
        items = response.json().get("message", {}).get("items", [])
    except Exception:
        return None

    if not items:
        return None

    item = items[0]
    candidate_title = " ".join(item.get("title") or [])
    if normalize_title(candidate_title) != normalize_title(title):
        return None

    if year:
        issued = item.get("issued", {}).get("date-parts", [[]])[0]
        candidate_year = issued[0] if issued else None
        if candidate_year and abs(int(candidate_year) - int(year)) > 1:
            return None

    return _normalize_doi(item.get("DOI"))


def parse_venue(venue):
    result = {"journal": venue or "", "volume": "", "issue": "", "pages": ""}
    if not venue:
        return result

    match = re.match(r"^(?P<journal>.*?)(?:\s+(?P<volume>\d+)(?:\s*\((?P<issue>[^)]+)\))?)?(?:,\s*(?P<pages>[\w-]+))?(?:,\s*\d{4})?$", venue)
    if match:
        groups = match.groupdict()
        result.update({key: value or "" for key, value in groups.items()})
        if not result["journal"]:
            result["journal"] = venue
    return result


def parse_year(pub, details):
    date_text = details.get("Publication date") or ""
    match = re.search(r"\d{4}", date_text or pub.get("year", ""))
    if match:
        return int(match.group(0))
    return None


def make_bibtex_key(title, year):
    words = re.findall(r"[A-Za-z0-9]+", title or "")
    stem = "".join(words[:5]) or "publication"
    return f"deng{year or ''}{stem[:60]}"


def make_bibtex(record):
    entry = {
        "ENTRYTYPE": "article",
        "ID": make_bibtex_key(record["标题"], record["年份"]),
        "title": _sanitize_bibtex_value(record["标题"]),
        "author": _sanitize_bibtex_value(record["作者"].replace(", ", " and ")),
        "journal": _sanitize_bibtex_value(record["期刊"]),
    }

    optional = {
        "year": record["年份"],
        "volume": record["卷号"],
        "number": record["期号"],
        "pages": record["页码"],
        "doi": record["DOI"],
        "url": record["链接"],
    }
    for key, value in optional.items():
        value = _sanitize_bibtex_value(value)
        if value:
            entry[key] = value
    return _entry_to_bibtex(entry)


def build_record(pub, details):
    venue_parts = parse_venue(pub.get("venue"))
    title = details.get("Title") or pub.get("title") or ""
    authors = details.get("Authors") or pub.get("authors") or ""
    year = parse_year(pub, details)
    doi = extract_doi_from_text(
        details.get("DOI"),
        details.get("Description"),
        pub.get("url"),
        " ".join(str(v) for v in details.values()),
    )
    if not doi:
        doi = crossref_doi(title, year)

    record = {
        "标题": title,
        "是否一作": is_first_author(authors),
        "是否通信": "未知（Scholar未提供）",
        "作者": authors,
        "期刊": details.get("Journal") or details.get("Conference") or venue_parts["journal"],
        "发表时间": details.get("Publication date") or str(year or pub.get("year") or ""),
        "年份": year or "",
        "卷号": details.get("Volume") or venue_parts["volume"],
        "期号": details.get("Issue") or venue_parts["issue"],
        "页码": details.get("Pages") or venue_parts["pages"],
        "DOI": doi or "",
        "引用数": pub.get("citations", 0),
        "链接": pub.get("url", ""),
        "BibTeX": "",
    }
    record["BibTeX"] = make_bibtex(record)
    return record


def write_excel(records, output_file):
    columns = ["标题", "是否一作", "是否通信", "作者", "期刊", "发表时间", "卷号", "期号", "页码", "DOI", "引用数", "链接", "BibTeX"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Google Scholar"
    ws.append(columns)

    for record in records:
        ws.append([record.get(column, "") for column in columns])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {
        "A": 56,
        "B": 10,
        "C": 22,
        "D": 42,
        "E": 34,
        "F": 16,
        "G": 10,
        "H": 10,
        "I": 16,
        "J": 28,
        "K": 10,
        "L": 44,
        "M": 70,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(output_file)


def main():
    service = GoogleScholarService(SCHOLAR_ID)
    publications = fetch_all_publications(service)
    print(f"Fetched {len(publications)} publication rows from Google Scholar")

    records = []
    for index, pub in enumerate(publications, start=1):
        try:
            details = fetch_detail(service, pub)
        except Exception as exc:
            print(f"[{index}/{len(publications)}] detail failed: {pub.get('title')} ({exc})")
            details = {}
        record = build_record(pub, details)
        records.append(record)
        print(f"[{index}/{len(publications)}] {record['标题']}")
        time.sleep(REQUEST_DELAY_SECONDS)

    records.sort(key=lambda item: (item.get("年份") or 0, item.get("引用数") or 0), reverse=True)
    write_excel(records, OUTPUT_FILE)
    print(f"Wrote {len(records)} rows to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
