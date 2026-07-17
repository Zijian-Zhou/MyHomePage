# -*- coding: utf-8 -*-
import re

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


AUTHOR_ID = "1682559"
AUTHOR_NAME = "Xiaoheng Deng"
OUTPUT_FILE = "dxc.xlsx"


def norm_name(value):
    value = re.sub(r"\([^)]*\)", "", value or "")
    value = re.sub(r"[^a-zA-Z\s]", " ", value).lower()
    return " ".join(value.split())


def name_variants(full_name):
    parts = norm_name(full_name).split()
    variants = {norm_name(full_name)}
    if len(parts) >= 2:
        first, last = parts[0], parts[-1]
        variants.update({f"{first[0]} {last}", f"{last} {first}", f"{last} {first[0]}"})
    return variants


PROFILE_VARIANTS = name_variants(AUTHOR_NAME)


def is_first_author(authors):
    if not authors:
        return "未知"
    first = authors[0]
    if str(first.get("authorId")) == AUTHOR_ID:
        return "是"
    return "是" if norm_name(first.get("name")) in PROFILE_VARIANTS else "否"


def bibtex_key(title, year):
    words = re.findall(r"[A-Za-z0-9]+", title or "")
    return "deng{}{}".format(year or "", ("".join(words[:5]) or "publication")[:60])


def clean_value(value):
    return str(value or "").replace("\r\n", " ").replace("\n", " ").strip()


def make_bibtex(record):
    fields = [
        ("title", record["标题"]),
        ("author", record["作者"].replace(", ", " and ")),
        ("journal", record["期刊"]),
        ("year", record["年份"]),
        ("volume", record["卷号"]),
        ("number", record["期号"]),
        ("pages", record["页码"]),
        ("doi", record["DOI"]),
        ("url", record["链接"]),
    ]
    lines = [f"@article{{{bibtex_key(record['标题'], record['年份'])},"]
    for key, value in fields:
        value = clean_value(value)
        if value:
            lines.append(f"  {key} = {{{value}}},")
    if len(lines) > 1:
        lines[-1] = lines[-1].rstrip(",")
    lines.append("}")
    return "\n".join(lines)


def fetch_papers():
    session = requests.Session()
    session.trust_env = False
    fields = ",".join([
        "title",
        "authors",
        "venue",
        "year",
        "publicationDate",
        "publicationTypes",
        "journal",
        "externalIds",
        "citationCount",
        "url",
        "publicationVenue",
    ])
    papers = []
    offset = 0
    while True:
        response = session.get(
            f"https://api.semanticscholar.org/graph/v1/author/{AUTHOR_ID}/papers",
            params={"limit": 100, "offset": offset, "fields": fields},
            timeout=60,
        )
        response.raise_for_status()
        payload = response.json()
        batch = payload.get("data", [])
        print(f"offset {offset}: {len(batch)}")
        papers.extend(batch)
        if not batch or "next" not in payload:
            break
        offset = payload["next"]
    return papers


def paper_to_record(paper):
    authors = paper.get("authors") or []
    journal = paper.get("journal") or {}
    external_ids = paper.get("externalIds") or {}
    doi = clean_value(external_ids.get("DOI")).lower()
    title = clean_value(paper.get("title"))
    venue = (
        clean_value(journal.get("name"))
        or clean_value(paper.get("venue"))
        or clean_value((paper.get("publicationVenue") or {}).get("name"))
    )
    record = {
        "标题": title,
        "是否一作": is_first_author(authors),
        "是否通信": "未知（数据源未提供）",
        "作者": ", ".join(clean_value(author.get("name")) for author in authors if author.get("name")),
        "期刊": venue,
        "发表时间": clean_value(paper.get("publicationDate")) or clean_value(paper.get("year")),
        "年份": paper.get("year") or "",
        "卷号": clean_value(journal.get("volume")),
        "期号": clean_value(journal.get("issue")),
        "页码": clean_value(journal.get("pages")),
        "DOI": doi,
        "引用数": paper.get("citationCount") if paper.get("citationCount") is not None else "",
        "链接": f"https://doi.org/{doi}" if doi else clean_value(paper.get("url")),
    }
    record["BibTeX"] = make_bibtex(record)
    return record


def write_excel(records):
    columns = ["标题", "是否一作", "是否通信", "作者", "期刊", "发表时间", "卷号", "期号", "页码", "DOI", "引用数", "链接", "BibTeX"]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Publications"
    sheet.append(columns)
    for record in records:
        sheet.append([record.get(column, "") for column in columns])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = [56, 10, 22, 42, 34, 16, 10, 10, 16, 30, 10, 44, 72]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = width

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(OUTPUT_FILE)


def main():
    seen = set()
    records = []
    for paper in fetch_papers():
        title_key = re.sub(r"[^a-z0-9]+", " ", clean_value(paper.get("title")).lower()).strip()
        if not title_key or title_key in seen:
            continue
        seen.add(title_key)
        records.append(paper_to_record(paper))
    records.sort(key=lambda item: (item.get("年份") or 0, item.get("引用数") or 0), reverse=True)
    write_excel(records)
    print(f"wrote {len(records)} rows to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
